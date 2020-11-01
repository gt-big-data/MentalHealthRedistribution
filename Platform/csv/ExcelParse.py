import openpyxl
import pandas as pd
from openpyxl import load_workbook

workbook = load_workbook(filename='IHME_USA_COUNTY_USE_INJ_MORTALITY_1980_2014_NATIONAL_Y2018M03D13.xlsx')
sheets = workbook.sheetnames

for i in range(0, len(sheets)):
    read_file = pd.read_excel('IHME_USA_COUNTY_USE_INJ_MORTALITY_1980_2014_NATIONAL_Y2018M03D13.xlsx', sheet_name = sheets[i])
    read_file.to_csv(sheets[i] + ".csv")